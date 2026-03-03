import openpyxl
import pandas as pd
import unicodedata
import os
from google_sheet import GoogleSheetManager  # 你的主程式檔名如果不是 google_sheet.py 就改掉
from datetime import datetime

TEMPLATE_XLSX = "templates/提品匯入範例.xlsx"
TARGET_SHEET = "上傳模板-非食品"
LAST_COL_LETTER = "AI"  # 你 Google Sheet 最後一欄（依你的 dic 是 AI）

# GoogleSheet欄位 -> 模板欄位（中文）
MAPPING = {
    "__分車類型__": "分車類型",
    "__品名__": "中文品名",
    "__品名__EN": "英文品名",
    
    "brand": "品牌",
    "origin_country": "生產國家",
    "barcode": "條碼",
    "tax_type": "應/免稅",
    "sale_price": "原價",

    "length_cm": "商品長度",
    "width_cm": "商品寬度",
    "height_cm": "商品高度",
    "weight_kg": "重量-淨重",

    "spec_name_1": "多規類型(一)",
    "spec_value_1": "多規類型(一)內容",
    "spec_name_2": "多規類型(二)",
    "spec_value_2": "多規類型(二)內容",

    "stock_qty": "活動庫存",
    "__是否庫控__": "是否庫控",

    "product_description": "商品完整說明",

    # 以下暫時寫死，因為 Google Sheet 沒有對應欄位，但模板需要
    "__商品分類__": "商品分類",
    "__商品規格__": "商品規格",
    "__重量毛重__": "重量-毛重",
    "__保存日期__": "保存日期",
    "__保存日期單位__": "保存日期單位",
    "__產品責任險__": "產品責任險",
    "__產品核准字號__": "產品核准字號",
    "__商品主圖1__": "商品主圖1",

    "__活動庫存開始日期__": "活動庫存開始日期",
    "__活動庫存結束日期__": "活動庫存結束日期",
}

# 這些欄位通常是「下拉選單 / 嚴格字串比對」
DROPDOWN_KEYS = {
    "tax_type",
    "spec_name_1",
    "spec_name_2",
    "__分車類型__",
    "__保存日期單位__",
    "__商品分類__",
}

def normalize_dropdown_text(v) -> str:
    s = "" if v is None else str(v)
    s = s.strip()
    # 把全形/半形/相容字形統一（解決 顏⾊ 這種）
    s = unicodedata.normalize("NFKC", s)
    # 把奇怪空白統一（有時是 NBSP）
    s = s.replace("\u00A0", " ")
    return s

def read_template_headers(ws, header_row: int = 1):
    """從 worksheet 讀 header 列，直到遇到空白欄位為止"""
    headers = []
    col = 1
    while True:
        v = ws.cell(row=header_row, column=col).value
        if v is None or str(v).strip() == "":
            break
        headers.append(str(v).strip())
        col += 1
    return headers

def calc_cart_type(row) -> str:
    """由 is_preorder + temperature_type 推導分車類型"""

    def norm_bool(v) -> bool:
        s = str(v).strip().lower()
        return s in {"1", "true", "yes", "y", "是", "勾選", "✓"}

    def norm_temp(v) -> str:
        s = str(v).strip()
        if s in {"冷藏", "冷凍"}:
            return "低溫"
        return "常溫"

    is_preorder = norm_bool(row.get("is_preorder", ""))
    temp = norm_temp(row.get("temperature_type", ""))

    prefix = "預購宅配" if is_preorder else "一般宅配"
    return prefix + temp

def calc_product_name(row) -> str:
    """中英品名：brand + sale_product_name"""
    brand = str(row.get("brand", "")).strip()
    sale_name = str(row.get("sale_product_name", "")).strip()

    if brand and sale_name:
        return f"{brand} {sale_name}".strip()
    return (brand or sale_name).strip()

def calc_fixed_fields(src_key, row):
    """目前先寫死測試值，未來 Google Sheet 建好欄位後只改這裡"""

    fixed_values = {
        "__商品分類__": "G200-Hair Care 洗髮用品",
        "__商品規格__": "1入",
        "__重量毛重__": 0.5,
        "__保存日期__": 365,
        "__保存日期單位__": "天",
        "__產品責任險__": "已投保產品責任險",
        "__產品核准字號__": "BSMI-TEST-001",
        "__商品主圖1__": "https://online.carrefour.com.tw/on/demandware.static/-/Sites-carrefour-tw-m-inner/default/images/scm/large/53563181/27230052351_67880800365462_1.jpg",
        "__活動庫存開始日期__": "2026-03-01",
        "__活動庫存結束日期__": "2026-12-31",
    }

    return fixed_values.get(src_key, "")

def export_keep_sheets_xlsx(
    df,
    template_xlsx: str,
    output_xlsx: str,
    target_sheet: str,
    mapping: dict,
    header_row: int = 1,
    data_start_row: int = 4,
    clear_down_to: int | None = None,
):
    """用模板檔當底，保留所有分頁與格式；只清 target_sheet 資料區再寫入 df"""
    wb = openpyxl.load_workbook(template_xlsx)

    if target_sheet not in wb.sheetnames:
        raise ValueError(f"找不到分頁：{target_sheet}，目前有：{wb.sheetnames}")
    ws = wb[target_sheet]

    # 讀模板 header（用來定位每個欄位在哪一欄）
    template_headers = read_template_headers(ws, header_row=header_row)
    if not template_headers:
        raise ValueError(f"{target_sheet} 第 {header_row} 列讀不到 header")

    header_to_col = {h: i + 1 for i, h in enumerate(template_headers)}
    max_col = len(template_headers)

    # 建立有效 pairs（只要模板有該欄位就收）
    pairs = []
    for src_key, tpl_col in mapping.items():
        if tpl_col in header_to_col:
            pairs.append((src_key, tpl_col))

    if not pairs:
        raise ValueError("mapping 沒有任何欄位成功對上（模板 header 對不到）")

    # 清空資料區
    if clear_down_to is None:
        clear_down_to = max(ws.max_row, data_start_row + len(df) + 50)

    for r in range(data_start_row, clear_down_to + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

    # 寫入資料
    r = data_start_row
    for _, row in df.iterrows():
        for src_key, tpl_col in pairs:

            # 1) 計算欄位
            if src_key == "__分車類型__":
                value = calc_cart_type(row)

            elif src_key in {"__品名__", "__品名__EN"}:
                value = calc_product_name(row)

            elif src_key == "__是否庫控__":
                value = "是"   # 你目前決定全部庫控

            # 2) 活動庫存必填：df 沒填就給測試值
            elif src_key == "stock_qty":
                v = str(row.get("stock_qty", "")).strip()
                value = v if v != "" else 10

            # 3) 其他固定欄位
            elif src_key.startswith("__"):
                value = calc_fixed_fields(src_key, row)

            # 4) 一般欄位
            else:
                value = row.get(src_key, "")

            # 5) 下拉欄位：做字形正規化（避免 顏⾊ 這種）
            if src_key in DROPDOWN_KEYS:
                value = normalize_dropdown_text(value)

            ws.cell(row=r, column=header_to_col[tpl_col]).value = value

        r += 1

    wb.save(output_xlsx)
    print("完成輸出（保留全部分頁）：", output_xlsx)

def ask_last_row(min_row: int) -> int:
    """讓使用者輸入要抓到第幾列（Google Sheet 的列號）"""
    s = input(f"請輸入要抓取到第幾列（>= {min_row}，例如 5）： ").strip()
    if s == "":
        return min_row + 1

    if not s.isdigit():
        raise ValueError("你輸入的不是數字，請輸入例如 5、100 這種列號。")

    last_row = int(s)
    if last_row < min_row:
        raise ValueError(f"最後一列不能小於資料起始列 {min_row}。")

    return last_row

if __name__ == "__main__":
    gs = GoogleSheetManager(sheet_config_file="config/google_sheet_config.json")

    # ✅ 建立 output 資料夾（如果不存在）
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    # ✅ 可選：自動加日期時間（避免覆蓋舊檔）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    last_row = ask_last_row(min_row=gs.data_start_row)

    end_cell = f"{LAST_COL_LETTER}{last_row}"
    print(f"\n將抓取範圍：{gs.sheet_name}!A{gs.data_start_row}:{end_cell}\n")

    df = gs.load_gs_data(end_cell=end_cell)
    print("抓到筆數：", len(df))

    # ✅ 存到 output 資料夾
    output_path = os.path.join(
        output_dir,
        f"提品匯入_非食品_{timestamp}.xlsx"
    )

    export_keep_sheets_xlsx(
        df=df,
        template_xlsx=TEMPLATE_XLSX,
        output_xlsx=output_path,
        target_sheet=TARGET_SHEET,
        mapping=MAPPING,
        header_row=1,
        data_start_row=4,
        clear_down_to=4 + len(df) + 50,
    )

    print("檔案已輸出到：", output_path)
